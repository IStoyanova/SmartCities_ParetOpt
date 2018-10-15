from Scenarios.Infrastructure import *
from openpyxl import Workbook
model1=Model("Electricity")
model2=Model("Gas")


T=96

#'Power to gas conversion: how much electric power will be supplied'
p2g_nom=p2g.maxPInput
conversionSchedule=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="Power2GasConverter")

#'Power to gas storage: how much gas power will be supplied'
p2g_sto_nom=storage.maxDischarge
storageSchedule=model1.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="Power2GasStorage")
storageSOC=model1.addVars(T+1,lb=0.1,ub=1.0,vtype=GRB.CONTINUOUS, name="Power2GasStorageSOC")

#'District heating CHP: how much electric power will be supplied'
chpDH_Schedule=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="dh_chp")

#'District Gas Source: how much gas power will be supplied'
gs_nom=100000000
gasSupplySchedule =model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="gas_source")



#'Schedule assigned to gas boilers of the buildings'
boil_nom=heater_boiler3.qNominal
boilerSchedule3=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler3")
boilerSchedule4=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler4")
boilerSchedule5=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler5")
boilerSchedule6=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler6")
boilerSchedule7=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler7")
boilerSchedule8=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler8")
boilerSchedule9=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler9")
boilerSchedule10=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler10")
boilerSchedule11=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler11")
boilerSchedule12=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler12")


#'Schedule assigned to heat pumps of the buildings'
heatFlow=[60]*192 #heat flow temperature 60C
hp_nom=heater_hp9.getNominalValues(heatFlow)
hpSchedule9 =model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp9")
hpSchedule10=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp10")
hpSchedule11=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp11")
hpSchedule12=model1.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp12")



#'Building gas import'
b3_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_gas")
b4_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_gas")
b5_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_gas")
b6_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_gas")
b7_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_gas")
b8_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_gas")
b9_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b9_gas")
b10_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b10_gas")
b11_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b11_gas")
b12_gas=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b12_gas")

#'Building heat import'
b3_th=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_th")
b4_th=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_th")
b5_th=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_th")
b6_th=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_th")
b7_th=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_th")
b8_th=model1.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_th")


model1.update()

#Heat balance for district heating region W
model1.addConstrs(chpDH_Schedule[t]*dh_CHP.qNominal==b3_th[t]+b4_th[t]+b5_th[t]+b6_th[t]+b7_th[t]+b8_th[t] for t in range(T))

#Heat balance for buildings W
model1.addConstrs(boilerSchedule3[t] * boil_nom + b3_th[t]== building3.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule4[t] * boil_nom + b4_th[t]== building4.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule5[t] * boil_nom + b5_th[t]== building5.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule6[t] * boil_nom + b6_th[t]== building6.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule7[t] * boil_nom + b7_th[t]== building7.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule8[t] * boil_nom + b8_th[t]== building8.get_power_curves()[1][t] for t in range(T))

#Heat balance for buildings E
model1.addConstrs(boilerSchedule9[t] * boil_nom + hpSchedule9[t] *hp_nom[1][t]== building9.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule10[t]* boil_nom + hpSchedule10[t]*hp_nom[1][t]== building10.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule11[t]* boil_nom + hpSchedule11[t]*hp_nom[1][t]== building11.get_power_curves()[1][t] for t in range(T))
model1.addConstrs(boilerSchedule12[t]* boil_nom + hpSchedule12[t]*hp_nom[1][t]== building12.get_power_curves()[1][t] for t in range(T))


#Storage balance storage units
model1.addConstr(storageSOC[0], '==', storage.ffInit)
model1.addConstrs((storageSOC[t + 1] == storageSOC[t] - storageSchedule[t] * p2g_sto_nom * storage.dT / storage.capacity for t in range(T)))



#Gas balance for buildings
model1.addConstrs((b3_gas[t] ==  boilerSchedule3[t]  * boil_nom/heater_boiler3.eta for t in range(T)))
model1.addConstrs((b4_gas[t] ==  boilerSchedule4[t]  * boil_nom/heater_boiler4.eta for t in range(T)))
model1.addConstrs((b5_gas[t] ==  boilerSchedule5[t]  * boil_nom/heater_boiler5.eta for t in range(T)))
model1.addConstrs((b6_gas[t] ==  boilerSchedule6[t]  * boil_nom/heater_boiler6.eta for t in range(T)))
model1.addConstrs((b7_gas[t] ==  boilerSchedule7[t]  * boil_nom/heater_boiler7.eta for t in range(T)))
model1.addConstrs((b8_gas[t] ==  boilerSchedule8[t]  * boil_nom/heater_boiler8.eta for t in range(T)))
model1.addConstrs((b9_gas[t] ==  boilerSchedule9[t]  * boil_nom/heater_boiler9.eta for t in range(T)))
model1.addConstrs((b10_gas[t] ==  boilerSchedule10[t]  * boil_nom/heater_boiler10.eta for t in range(T)))
model1.addConstrs((b11_gas[t] ==  boilerSchedule11[t]  * boil_nom/heater_boiler11.eta for t in range(T)))
model1.addConstrs((b12_gas[t] ==  boilerSchedule12[t]  * boil_nom/heater_boiler12.eta for t in range(T)))


# Total gas supply balance
model1.addConstrs((gasSupplySchedule[t] * gs_nom +
                conversionSchedule[t] * p2g_nom * p2g.eta +
                storageSchedule[t] * p2g_sto_nom
                ==
                chpDH_Schedule[t] * dh_CHP.fNominal +
                b3_gas[t] +
                b4_gas[t] +
                b5_gas[t] +
                b6_gas[t] +
                b7_gas[t] +
                b8_gas[t] +
                b9_gas[t] +
                b10_gas[t] +
                b11_gas[t] +
                b12_gas[t]
                for t in range(T)))



model1.setObjective(sum((chpDH_Schedule[t] * dh_CHP.fNominal +
                        b3_gas[t] +
                        b4_gas[t] +
                        b5_gas[t] +
                        b6_gas[t] +
                        b7_gas[t] +
                        b8_gas[t] +
                        b9_gas[t] +
                        b10_gas[t] +
                        b11_gas[t] +
                        b12_gas[t]) * 0.2*900 / 3600 / 1000
                       for t in range(T)), GRB.MINIMIZE)


print('Optimization of the gas domain')
model1.optimize()
print()




#'Hydroelectric pump storage: how much electric power will be supplied'
pump_nom=pump.maxDischarge
pumpSchedule=model2.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="heps")
pumpSOC=model2.addVars(T+1,lb=0.05,ub=1.0,vtype=GRB.CONTINUOUS, name="pumpSOC")

#'District Gas Source: how much water power will be supplied'
ws_nom=100000000
waterSupplySchedule =model2.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="water_source")

#'Battery schedule assigned'
bat_nom=battery9.maxDischarge
batterySchedule9=model2.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery9")
batterySchedule10=model2.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery10")
batterySchedule11=model2.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery11")
batterySchedule12=model2.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery12")

batterySOC9=model2.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC9")
batterySOC10=model2.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC10")
batterySOC11=model2.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC11")
batterySOC12=model2.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC12")

imported=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS,name='imported')
utilization_pv=model2.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS,name='utilization_pv')
utilization_wec=model2.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS,name='utilization_wec')

#'Building electricity import'
b3_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_ele")
b4_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_ele")
b5_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_ele")
b6_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_ele")
b7_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_ele")
b8_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_ele")
b9_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b9_ele")
b10_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b10_ele")
b11_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b11_ele")
b12_ele=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b12_ele")

#'Building water import'
b3_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_wat")
b4_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_wat")
b5_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_wat")
b6_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_wat")
b7_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_wat")
b8_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_wat")
b9_wat =model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b9_wat")
b10_wat=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b10_wat")
b11_wat=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b11_wat")
b12_wat=model2.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b12_wat")

# Power balance for buildings W
model2.addConstrs(b3_ele[t] == building3.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b4_ele[t] == building4.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b5_ele[t] == building5.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b6_ele[t] == building6.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b7_ele[t] == building7.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b8_ele[t] == building8.get_power_curves()[0][t] for t in range(T))

# Power balance for buildings E
model2.addConstrs(b9_ele[t] + batterySchedule9[t] * bat_nom - hpSchedule9[t].X * hp_nom[0][t]== building9.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b10_ele[t] + batterySchedule10[t] * bat_nom - hpSchedule10[t].X * hp_nom[0][t]== building10.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b11_ele[t] + batterySchedule11[t] * bat_nom - hpSchedule11[t].X * hp_nom[0][t]== building11.get_power_curves()[0][t] for t in range(T))
model2.addConstrs(b12_ele[t] + batterySchedule12[t] * bat_nom - hpSchedule12[t].X * hp_nom[0][t]== building12.get_power_curves()[0][t] for t in range(T))

#SOC balance of electrical storage devices
model2.addConstr(batterySOC9[0], '==', battery9.socInit)
model2.addConstr(batterySOC10[0], '==', battery10.socInit)
model2.addConstr(batterySOC11[0], '==', battery11.socInit)
model2.addConstr(batterySOC12[0], '==', battery12.socInit)
model2.addConstr(pumpSOC[0], '==', pump.ffInit)

model2.addConstrs((batterySOC9[t+1]  == batterySOC9[t]  - batterySchedule9[t]  *bat_nom *battery9.dT /battery9.capacity  for t in range(T)))
model2.addConstrs((batterySOC10[t+1] == batterySOC10[t] - batterySchedule10[t] *bat_nom *battery10.dT/battery10.capacity for t in range(T)))
model2.addConstrs((batterySOC11[t+1] == batterySOC11[t] - batterySchedule11[t] *bat_nom *battery11.dT/battery11.capacity for t in range(T)))
model2.addConstrs((batterySOC12[t+1] == batterySOC12[t] - batterySchedule12[t] *bat_nom *battery12.dT/battery12.capacity for t in range(T)))
model2.addConstrs((pumpSOC[t + 1] == pumpSOC[t] - pumpSchedule[t] * pump_nom * pump.dT / pump.energyCapacity for t in range(T)))

# Total power balance
model2.addConstrs((utilization_pv[t]*cityDistrict.getPVPower()[t] +
                   utilization_wec[t] *cityDistrict.getWindEnergyConverterPower()[t] +
                  chpDH_Schedule[t].X * dh_CHP.pNominal +
                  pumpSchedule[t] * pump_nom+
                  imported[t]
                ==
                conversionSchedule[t].X * p2g_nom +
                b3_ele[t] +
                b4_ele[t] +
                b5_ele[t] +
                b6_ele[t] +
                b7_ele[t] +
                b8_ele[t] +
                b9_ele[t] +
                b10_ele[t] +
                b11_ele[t] +
                b12_ele[t] for t in range(T)))

#Water balance for buildings
model2.addConstrs((b3_wat[t] ==  building3.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b4_wat[t] ==  building4.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b5_wat[t] ==  building5.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b6_wat[t] ==  building6.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b7_wat[t] ==  building7.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b8_wat[t] ==  building8.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b9_wat[t] ==  building9.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b10_wat[t] ==  building10.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b11_wat[t] ==  building11.get_waterDemand_profile()[t] for t in range(T)))
model2.addConstrs((b12_wat[t] ==  building12.get_waterDemand_profile()[t] for t in range(T)))


#Total water balance
model2.addConstrs((waterSupplySchedule[t] * ws_nom+
                  pumpSchedule[t]*pump_nom/(pump.d*pump.g*pump.h)
                  ==
                  b3_wat[t] +
                  b4_wat[t] +
                  b5_wat[t] +
                  b6_wat[t] +
                  b7_wat[t] +
                  b8_wat[t] +
                  b9_wat[t] +
                  b10_wat[t] +
                  b11_wat[t] +
                  b12_wat[t]
                  for t in range(T)))

model2.setObjective(sum((imported[t]/0.4) * 0.2*900 / 3600 / 1000
                       for t in range(T)), GRB.MINIMIZE)


print('Optimization of the electricity-water domain')
model2.optimize()
print()

e=2
wb = Workbook()
ws = wb.active



ws.cell(row=8, column=e + 2, value='co2pumpshedule')
ws.cell(row=8 + (T + 3) , column=e + 2, value='co2pumpgasshedule')
ws.cell(row=8 + (T + 3)*2   , column=e + 2, value='generalHP    ')
ws.cell(row=8 + (T + 3) *3, column=e + 2, value='conversionSchedule')
ws.cell(row=8 + (T + 3) *4, column=e + 2, value='storageSchedule')
ws.cell(row=8 + (T + 3) *5, column=e + 2, value='storageSOC')
ws.cell(row=8 + (T + 3) *6, column=e + 2, value='pumpSchedule')
ws.cell(row=8 + (T + 3) *7, column=e + 2, value='pumpSOC')
ws.cell(row=8 + (T + 3) *8, column=e + 2, value='chpDH_Schedule')
ws.cell(row=8 + (T + 3) *9, column=e + 2, value='gasSupplySchedule')
ws.cell(row=8 + (T + 3) *10, column=e + 2, value='  waterSupplySchedule  ')
ws.cell(row=8 + (T + 3) *11, column=e + 2, value='  boilerSchedule3  ')
ws.cell(row=8 + (T + 3) *12, column=e + 2, value='  hpSchedule9  ')
ws.cell(row=8 + (T + 3) *13, column=e + 2, value=' batterySchedule9   ')
ws.cell(row=8 + (T + 3) *14, column=e + 2, value='  batterySOC9  ')
ws.cell(row=8 + (T + 3) * 15, column=e + 2, value='   imported ')
ws.cell(row=8 + (T + 3) * 16, column=e + 2, value='  utilization_pv  ')
ws.cell(row=8 + (T + 3) * 17, column=e + 2, value='  utilization_wec  ')
ws.cell(row=8 + (T + 3) * 18, column=e + 2, value='  b3_ele  ')
ws.cell(row=8 + (T + 3) * 19, column=e + 2, value='  b3_gas  ')
ws.cell(row=8 + (T + 3) * 20, column=e + 2, value=' b3_th   ')
ws.cell(row=8 + (T + 3) * 21, column=e + 2, value=' b3_wat   ')

for t in range(T):
    ws.cell(row=9 + (T + 3) * 3 + t, column=e + 2, value=conversionSchedule[t].X)
    ws.cell(row=9 + (T + 3) * 4 + t, column=e + 2, value=storageSchedule[t].X)
    ws.cell(row=9 + (T + 3) * 5 + t, column=e + 2, value=storageSOC[t].X)
    ws.cell(row=9 + (T + 3) * 6 + t, column=e + 2, value=pumpSchedule[t].X)
    ws.cell(row=9 + (T + 3) * 7 + t, column=e + 2, value=pumpSOC[t].X)
    ws.cell(row=9 + (T + 3) * 8 + t, column=e + 2, value=chpDH_Schedule[t].X)
    ws.cell(row=9 + (T + 3) * 9 + t, column=e + 2, value=gasSupplySchedule[t].X)
    ws.cell(row=9 + (T + 3) * 10 + t, column=e + 2, value=waterSupplySchedule[t].X)
    ws.cell(row=9 + (T + 3) * 11 + t, column=e + 2, value=boilerSchedule3[t].X)
    ws.cell(row=9 + (T + 3) * 12 + t, column=e + 2, value=hpSchedule9[t].X)
    ws.cell(row=9 + (T + 3) * 13 + t, column=e + 2, value=batterySchedule9[t].X)
    ws.cell(row=9 + (T + 3) * 14 + t, column=e + 2, value=batterySOC9[t].X)
    ws.cell(row=9 + (T + 3) * 15 + t, column=e + 2, value=imported[t].X)
    ws.cell(row=9 + (T + 3) * 16 + t, column=e + 2, value=utilization_pv[t].X)
    ws.cell(row=9 + (T + 3) * 17 + t, column=e + 2, value=utilization_wec[t].X)
    ws.cell(row=9 + (T + 3) * 18 + t, column=e + 2, value=b3_ele[t].X)
    ws.cell(row=9 + (T + 3) * 19 + t, column=e + 2, value=b3_gas[t].X)
    ws.cell(row=9 + (T + 3) * 20 + t, column=e + 2, value=b3_th[t].X)
    ws.cell(row=9 + (T + 3) * 21 + t, column=e + 2, value=b3_wat[t].X)







wb.save('optimization.xlsx')




if __name__ == '__main__':
    imported_Ele=0
    nonUtil_Ele=0
    directEmission=0
    for i in range(T):
        imported_Ele+=imported[i].X/1000/3600*900
        nonUtil_Ele+=((1 - utilization_wec[i].X) * cityDistrict.getWindEnergyConverterPower()[i] + (1 - utilization_pv[i].X) *cityDistrict.getPVPower()[i])/1000/3600*900
        directEmission+=(chpDH_Schedule[i].X * dh_CHP.fNominal +
                            b3_gas[i].X +
                            b4_gas[i].X +
                            b5_gas[i].X +
                            b6_gas[i].X +
                            b7_gas[i].X +
                            b8_gas[i].X +
                            b9_gas[i].X +
                            b10_gas[i].X +
                            b11_gas[i].X +
                            b12_gas[i].X) * 0.2/1000/3600*900
    print('import total',imported_Ele, 'kWh')
    print('non-utilized total',nonUtil_Ele,'kWh')
    print('direct emission',directEmission, 'kg CO2')

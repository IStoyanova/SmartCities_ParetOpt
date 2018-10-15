"""file modified by ferran after coments where it is mentioned. this file has been highly modified in the constraints and objectives parts"""
from Scenarios.Infrastructure import *
from openpyxl import Workbook

print('multiobjective Optimization of the multi-domain operation')
print()

"""lines modifed"""
#vehicle curves
ca=bus1.get_energy_curves()
cb=bus1.get_charging_curves()
cc=EVcomm.get_vehicle_curves()[0]
cd=EVcomm.get_vehicle_curves()[1]
ch=ngcpleass.get_vehicle_curves()[0]
ci=ngcpleass.get_vehicle_curves()[1]

#shop curves
ce=bakery.get_total_energy_curves()
cf=business.get_total_energy_curves()
cg=shop.get_total_energy_curves()
"""lines modifed till here"""

model=Model("OptimumOperation")




T=96
"""lines modifed"""
dummy=model.addVar(vtype=GRB.BINARY)

variationele=model.addVar(lb=0.0,vtype=GRB.CONTINUOUS, name="variationelectrcity")
variationgas=model.addVar(lb=0.0,vtype=GRB.CONTINUOUS, name="variationgas")

#bus energy demand and fuel tank level

busfill=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="buspump")
bustankSOC=model.addVars(T+1,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="busSOC")

#EV energy demand and battery level

EVcharge=model.addVars(T,lb=0.0,ub=0.30,vtype=GRB.CONTINUOUS, name="electric battery charge")
EVbattSOC=model.addVars(T+1,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="electric cars SOC")

#ngc energy demand and fuel tank level

ngcfill=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="natural gas car pump")
ngctankSOC=model.addVars(T+1,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="ngc SOC")


#'CO2 capturing and sequestration'
co2max=co2capseq.maxinput
co2pumpshedule=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="co2pump")
co2pumpgasshedule=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="co2pumpfrom gas")


#'electric heating (general heat pump) for buildings 3 to 8' (not used)
hpG_nom=600000
generalHP=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hpG")
"""lines modifed till here"""
#'Power to gas conversion: how much electric power will be supplied'
p2g_nom=p2g.maxPInput
conversionSchedule=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="Power2GasConverter")

#'Power to gas storage: how much gas power will be supplied'
p2g_sto_nom=storage.maxDischarge
storageSchedule=model.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="Power2GasStorage")
storageSOC=model.addVars(T+1,lb=0.1,ub=1.0,vtype=GRB.CONTINUOUS, name="Power2GasStorageSOC")

#'Hydroelectric pump storage: how much electric power will be supplied'
pump_nom=pump.maxDischarge
pumpSchedule=model.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="heps")
pumpSOC=model.addVars(T+1,lb=0.05,ub=1.0,vtype=GRB.CONTINUOUS, name="pumpSOC")

#'District heating CHP: how much electric power will be supplied'
chpDH_Schedule=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="dh_chp")

#'District Gas Source: how much gas power will be supplied'
gs_nom=100000000
gasSupplySchedule =model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="gas_source")

#'District Gas Source: how much water power will be supplied'
ws_nom=100000000
waterSupplySchedule =model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="water_source")

#'Schedule assigned to gas boilers of the buildings'
boil_nom=heater_boiler3.qNominal
boilerSchedule3=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler3")
boilerSchedule4=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler4")
boilerSchedule5=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler5")
boilerSchedule6=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler6")
boilerSchedule7=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler7")
boilerSchedule8=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler8")
boilerSchedule9=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler9")
boilerSchedule10=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler10")
boilerSchedule11=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler11")
boilerSchedule12=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="boiler12")


#'Schedule assigned to heat pumps of the buildings'
heatFlow=[60]*192 #heat flow temperature 60C
hp_nom=heater_hp9.getNominalValues(heatFlow)
hpSchedule9 =model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp9")
hpSchedule10=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp10")
hpSchedule11=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp11")
hpSchedule12=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS, name="hp12")

#'Battery schedule assigned'
bat_nom=battery9.maxDischarge
batterySchedule9=model.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery9")
batterySchedule10=model.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery10")
batterySchedule11=model.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery11")
batterySchedule12=model.addVars(T,lb=-1.0,ub=1.0,vtype=GRB.CONTINUOUS, name="battery12")

batterySOC9=model.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC9")
batterySOC10=model.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC10")
batterySOC11=model.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC11")
batterySOC12=model.addVars(T+1,lb=0.2,ub=1.0,vtype=GRB.CONTINUOUS, name="batterySOC12")

imported=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS,name='imported')
utilization_pv=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS,name='utilization_pv')
utilization_wec=model.addVars(T,lb=0.0,ub=1.0,vtype=GRB.CONTINUOUS,name='utilization_wec')


#'Building electricity import'
b3_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_ele")
b4_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_ele")
b5_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_ele")
b6_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_ele")
b7_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_ele")
b8_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_ele")
b9_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b9_ele")
b10_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b10_ele")
b11_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b11_ele")
b12_ele=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b12_ele")

#'Building gas import'
b3_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_gas")
b4_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_gas")
b5_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_gas")
b6_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_gas")
b7_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_gas")
b8_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_gas")
b9_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b9_gas")
b10_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b10_gas")
b11_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b11_gas")
b12_gas=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b12_gas")

#'Building heat import'
b3_th=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_th")
b4_th=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_th")
b5_th=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_th")
b6_th=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_th")
b7_th=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_th")
b8_th=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_th")


#'Building water import'
b3_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b3_wat")
b4_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b4_wat")
b5_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b5_wat")
b6_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b6_wat")
b7_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b7_wat")
b8_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b8_wat")
b9_wat =model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b9_wat")
b10_wat=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b10_wat")
b11_wat=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b11_wat")
b12_wat=model.addVars(T,lb=0.0,vtype=GRB.CONTINUOUS, name="b12_wat")

model.update()







"""lines modifed"""
#co2 max elimination equal to 80 percent
model.addConstrs(co2max*co2capseq.rate*(co2pumpgasshedule[t]+co2pumpshedule[t])<=co2capseq.maxelimination*(chpDH_Schedule[t] * dh_CHP.fNominal +
                        b3_gas[t] +
                        b4_gas[t] +
                        b5_gas[t] +
                        b6_gas[t] +
                        b7_gas[t] +
                        b8_gas[t] +
                        b9_gas[t] +
                        b10_gas[t] +
                        b11_gas[t] +
                        b12_gas[t]) for t in range(T))



#Heat balance for district heating region W
model.addConstrs(chpDH_Schedule[t]*dh_CHP.qNominal+hpG_nom*0.9*generalHP[t]==b3_th[t]+b4_th[t]+b5_th[t]+b6_th[t]+b7_th[t]+b8_th[t] for t in range(T))

#Heat balance for buildings W
model.addConstrs(boilerSchedule3[t] * boil_nom + b3_th[t]== building3.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule4[t] * boil_nom + b4_th[t]== building4.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule5[t] * boil_nom + b5_th[t]== building5.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule6[t] * boil_nom + b6_th[t]== building6.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule7[t] * boil_nom + b7_th[t]== building7.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule8[t] * boil_nom + b8_th[t]== building8.get_power_curves()[1][t] for t in range(T))

#Heat balance for buildings E
model.addConstrs(boilerSchedule9[t] * boil_nom + hpSchedule9[t] *hp_nom[1][t]== building9.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule10[t]* boil_nom + hpSchedule10[t]*hp_nom[1][t]== building10.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule11[t]* boil_nom + hpSchedule11[t]*hp_nom[1][t]== building11.get_power_curves()[1][t] for t in range(T))
model.addConstrs(boilerSchedule12[t]* boil_nom + hpSchedule12[t]*hp_nom[1][t]== building12.get_power_curves()[1][t] for t in range(T))

# Power balance for buildings W
model.addConstrs(b3_ele[t] == building3.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b4_ele[t] == building4.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b5_ele[t] == building5.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b6_ele[t] == building6.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b7_ele[t] == building7.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b8_ele[t] == building8.get_power_curves()[0][t] for t in range(T))

# Power balance for buildings E
model.addConstrs(b9_ele[t] + batterySchedule9[t] * bat_nom - hpSchedule9[t] * hp_nom[0][t]== building9.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b10_ele[t] + batterySchedule10[t] * bat_nom - hpSchedule10[t] * hp_nom[0][t]== building10.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b11_ele[t] + batterySchedule11[t] * bat_nom - hpSchedule11[t] * hp_nom[0][t]== building11.get_power_curves()[0][t] for t in range(T))
model.addConstrs(b12_ele[t] + batterySchedule12[t] * bat_nom - hpSchedule12[t] * hp_nom[0][t]== building12.get_power_curves()[0][t] for t in range(T))

#Storage balance storage units
model.addConstr(batterySOC9[0], '==', battery9.socInit)
model.addConstr(batterySOC10[0], '==', battery10.socInit)
model.addConstr(batterySOC11[0], '==', battery11.socInit)
model.addConstr(batterySOC12[0], '==', battery12.socInit)
model.addConstr(storageSOC[0], '==', storage.ffInit)
model.addConstr(pumpSOC[0], '==', pump.ffInit)


model.addConstrs((batterySOC9[t+1]  == batterySOC9[t]  - batterySchedule9[t]  *bat_nom *battery9.dT /battery9.capacity  for t in range(T)))
model.addConstrs((batterySOC10[t+1] == batterySOC10[t] - batterySchedule10[t] *bat_nom *battery10.dT/battery10.capacity for t in range(T)))
model.addConstrs((batterySOC11[t+1] == batterySOC11[t] - batterySchedule11[t] *bat_nom *battery11.dT/battery11.capacity for t in range(T)))
model.addConstrs((batterySOC12[t+1] == batterySOC12[t] - batterySchedule12[t] *bat_nom *battery12.dT/battery12.capacity for t in range(T)))
model.addConstrs((storageSOC[t + 1] == storageSOC[t] - storageSchedule[t] * p2g_sto_nom * storage.dT / storage.capacity for t in range(T)))
model.addConstrs((pumpSOC[t + 1] == pumpSOC[t] - pumpSchedule[t] * pump_nom * pump.dT / pump.energyCapacity for t in range(T)))

# Total ele power balance
model.addConstrs((utilization_pv[t]*cityDistrict.getPVPower()[t] +
                  utilization_wec[t] *cityDistrict.getWindEnergyConverterPower()[t] +
                  chpDH_Schedule[t] * dh_CHP.pNominal +
                  pumpSchedule[t] * pump_nom+
                  imported[t]
                ==
                EVcharge[t]*EVcomm.get_total_energy_vehicles()+

                conversionSchedule[t] * p2g_nom +
                b3_ele[t] +
                b4_ele[t] +
                b5_ele[t] +
                b6_ele[t] +
                b7_ele[t] +
                b8_ele[t] +
                b9_ele[t] +
                b10_ele[t] +
                b11_ele[t] +
                b12_ele[t]+
                  ce[t] + cf[t] + cg[t] + co2max * co2pumpshedule[t]+ hpG_nom*generalHP[t] for t in range(T)))


"""   """


#Gas balance for buildings
model.addConstrs((b3_gas[t] ==  boilerSchedule3[t]  * boil_nom/heater_boiler3.eta for t in range(T)))
model.addConstrs((b4_gas[t] ==  boilerSchedule4[t]  * boil_nom/heater_boiler4.eta for t in range(T)))
model.addConstrs((b5_gas[t] ==  boilerSchedule5[t]  * boil_nom/heater_boiler5.eta for t in range(T)))
model.addConstrs((b6_gas[t] ==  boilerSchedule6[t]  * boil_nom/heater_boiler6.eta for t in range(T)))
model.addConstrs((b7_gas[t] ==  boilerSchedule7[t]  * boil_nom/heater_boiler7.eta for t in range(T)))
model.addConstrs((b8_gas[t] ==  boilerSchedule8[t]  * boil_nom/heater_boiler8.eta for t in range(T)))
model.addConstrs((b9_gas[t] ==  boilerSchedule9[t]  * boil_nom/heater_boiler9.eta for t in range(T)))
model.addConstrs((b10_gas[t] ==  boilerSchedule10[t]  * boil_nom/heater_boiler10.eta for t in range(T)))
model.addConstrs((b11_gas[t] ==  boilerSchedule11[t]  * boil_nom/heater_boiler11.eta for t in range(T)))
model.addConstrs((b12_gas[t] ==  boilerSchedule12[t]  * boil_nom/heater_boiler12.eta for t in range(T)))


# Total gas supply balance
model.addConstrs((gasSupplySchedule[t] * gs_nom +
                conversionSchedule[t] * p2g_nom * p2g.eta +
                storageSchedule[t] * p2g_sto_nom
                  ==
                busfill[t]*bus1.get_total_energy_vehicles() +
                ngcfill[t]*ngcpleass.get_total_energy_vehicles() +
                chpDH_Schedule[t] * dh_CHP.fNominal +
                b3_gas[t] +
                b4_gas[t] +
                b5_gas[t] +
                b6_gas[t] +
                b7_gas[t] +
                b8_gas[t] +
                b9_gas[t] +
                b10_gas[t] +
                b11_gas[t] +
                b12_gas[t]+ co2max * co2pumpgasshedule[t]
                for t in range(T)))



#bus working system

model.addConstr(bustankSOC[0]== 0)
model.addConstrs(  bustankSOC[t+1] == bustankSOC[t] - ca[t] + busfill[t] for t in range(T))
model.addConstrs(   bustankSOC[t]+ busfill[t] <= 1 for t in range(T))
model.addConstrs(    busfill[t] <= cb[t] for t in range(T))




#ngc working system

model.addConstr(ngctankSOC[0]== 0)
model.addConstrs(ngctankSOC[t+1] == ngctankSOC[t] - ch[t] + ngcfill[t] for t in range(T))
model.addConstrs(ngctankSOC[t]+ ngcfill[t] <= 1 for t in range(T))
model.addConstrs(ngcfill[t] <= ci[t] for t in range(T))


#EV working system
model.addConstr(  EVbattSOC[0], '==', EVbattSOC[95])
model.addConstrs(   EVbattSOC[t]+ EVcharge[t] <= 1 for t in range(T))
model.addConstrs(  EVbattSOC[t+1] == EVbattSOC[t] - cc[t] +EVcharge[t] for t in range(T))
model.addConstrs(  EVcharge[t] <= cd[t] for t in range(T))



#Water balance for buildings
model.addConstrs((b3_wat[t] ==  building3.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b4_wat[t] ==  building4.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b5_wat[t] ==  building5.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b6_wat[t] ==  building6.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b7_wat[t] ==  building7.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b8_wat[t] ==  building8.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b9_wat[t] ==  building9.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b10_wat[t] ==  building10.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b11_wat[t] ==  building11.get_waterDemand_profile()[t] for t in range(T)))
model.addConstrs((b12_wat[t] ==  building12.get_waterDemand_profile()[t] for t in range(T)))


#Total water balance
model.addConstrs((waterSupplySchedule[t] * ws_nom+
                  pumpSchedule[t]*pump_nom/(pump.d*pump.g*pump.h)
                  ==
                  b3_wat[t] +
                  b4_wat[t] +
                  b5_wat[t] +
                  b6_wat[t] +
                  b7_wat[t] +
                  b8_wat[t] +
                  b9_wat[t] +
                  b10_wat[t] +
                  b11_wat[t] +
                  b12_wat[t]
                  for t in range(T)))


#uncoment to flaten the electical grid imported energy
model.addConstr(imported[0], '==', imported[95])
model.addConstrs(((imported[t]-imported[t+1])*900 / 3600 / 1000  <= variationele  for t in range(T-1)))
model.addConstrs(((imported[t]-imported[t+1])*900 / 3600 / 1000  >= -variationele  for t in range(T-1)))


#uncoment to flaten the gas extraction curve
model.addConstr(gasSupplySchedule[0], '==', gasSupplySchedule[95])
model.addConstrs(((gasSupplySchedule[t]-gasSupplySchedule[t+1])* gs_nom*900 / 3600 / 1000  <= variationgas  for t in range(T-1)))
model.addConstrs(((gasSupplySchedule[t]-gasSupplySchedule[t+1])* gs_nom *900 / 3600 / 1000 >= -variationgas  for t in range(T-1)))


#shuting down interfaces or variables
model.addConstrs(generalHP[t]== 0       for t in range(T))
#model.addConstrs(conversionSchedule[t]== 0       for t in range(T))
#model.addConstrs(storageSchedule[t]== 0       for t in range(T))
#model.addConstrs(pumpSchedule[t]== 0       for t in range(T))
#model.addConstrs(chpDH_Schedule[t]== 0       for t in range(T))
#model.addConstrs(gasSupplySchedule[t]== 0       for t in range(T))
#model.addConstrs(boilerSchedule3[t]== 0       for t in range(T))
#model.addConstrs(hpSchedule9[t]== 0       for t in range(T))
#model.addConstrs(batterySchedule9[t]== 0       for t in range(T))
model.addConstrs(co2pumpshedule[t]== 0       for t in range(T))
model.addConstrs(co2pumpgasshedule[t]== 0       for t in range(T))
#model.addConstrs(busfill[t]== 0       for t in range(T))
#model.addConstrs([t]== 0       for t in range(T))
#model.addConstrs([t]== 0       for t in range(T))
#model.addConstrs([t]== 0       for t in range(T))











# model.setObjectiveN(formula,index(atarts at 0),priority=1, abstol=0)
#seting objectives: to optimze the objective uncoment and set the index of the objectives you need (dont repeat),
#set the piority allso in order from highest the most importat objective to lowest the least important they should not
#and the lowest number should be 1

# inedxes of priority automatically filling priority and index, results will be shown in the order of priority
indvarele=2
indvargas=4
indco2=5
indimpo=3
indgascup=1
indEVvahrge=0


model.setObjectiveN(variationele,indvarele,priority=7-indvarele, abstol=0)

model.setObjectiveN(variationgas,indvargas,priority=7-indvargas, abstol=0)

model.setObjectiveN(sum((imported[t]/0.4+
                        chpDH_Schedule[t] * dh_CHP.fNominal +
                        b3_gas[t] +
                        b4_gas[t] +
                        b5_gas[t] +
                        b6_gas[t] +
                        b7_gas[t] +
                        b8_gas[t] +
                        b9_gas[t] +
                        b10_gas[t] +
                        b11_gas[t] +
                        b12_gas[t]-
                        (co2max * (co2pumpgasshedule[t]+co2pumpshedule[t])*3)) *0.2*900 / 3600 / 1000
                       for t in range(T)),indco2,priority=7-indco2, abstol=0)

model.setObjectiveN(sum((imported[t]) * 900 / 3600 / 1000
                       for t in range(T)),indimpo,priority=7-indimpo, abstol=0)

model.setObjectiveN(sum((gasSupplySchedule[t] * gs_nom) * 900 / 3600 / 1000
                       for t in range(T)),indgascup,priority=7-indgascup, abstol=0)

model.setObjectiveN((-1)*sum(EVbattSOC[t]
                       for t in range(T)),indEVvahrge,priority=7-indEVvahrge, abstol=0)

#optimization of CO2 emisisons
"""model.setObjectiveN(sum((imported[t]/0.4+
                        chpDH_Schedule[t] * dh_CHP.fNominal +
                        b3_gas[t] +
                        b4_gas[t] +
                        b5_gas[t] +
                        b6_gas[t] +
                        b7_gas[t] +
                        b8_gas[t] +
                        b9_gas[t] +
                        b10_gas[t] +
                        b11_gas[t] +
                        b12_gas[t]-
                        (co2max * (co2pumpgasshedule[t]+co2pumpshedule[t])*3)) *0.2*900 / 3600 / 1000
                       for t in range(T)),0,priority=1, abstol=0)"""



#optimization of gas suply
"""model.setObjectiveN(sum((gasSupplySchedule[t] * gs_nom) * 900 / 3600 / 1000
                       for t in range(T)),0,priority=5, abstol=0)"""

"""model.setObjectiveN(sum((chpDH_Schedule[t] * dh_CHP.fNominal +
                        b3_gas[t] +
                        b4_gas[t] +
                        b5_gas[t] +
                        b6_gas[t] +
                        b7_gas[t] +
                        b8_gas[t] +
                        b9_gas[t] +
                        b10_gas[t] +
                        b11_gas[t] +
                        b12_gas[t]) * 0.2*900 / 3600 / 1000
                       for t in range(T)), 0,priority=1, abstol=0)"""


#optimization of electrical energy imported from the grid
"""model.setObjectiveN(sum((imported[t]) * 900 / 3600 / 1000
                       for t in range(T)),1,priority=4, abstol=0)"""

"""model.setObjectiveN(sum((imported[t]/0.4) * 0.2*900 / 3600 / 1000
                       for t in range(T)),1,priority=2, abstol=0)"""

#optimization of fuel in bus tank
"""model.setObjectiveN(sum(bustankSOC[t]
                       for t in range(T)),2,priority=3, abstol=0)"""

#optimization of max SOC in EV batteries
"""model.setObjectiveN((-1)*sum(EVbattSOC[t]
                       for t in range(T)),3,priority=2, abstol=0)"""


#optimization of final state of storage units
"""model.setObjectiveN((batterySOC9[95]+batterySOC10[95]+batterySOC11[95]+batterySOC12[95]+storageSOC[95]+pumpSOC[95])*-1,
                        2,priority=1, abstol=0)"""

#optimization of CO2 emizzions of electric generating objects
"""model.setObjectiveN(sum((imported[t]/0.4+
                        chpDH_Schedule[t] * dh_CHP.fNominal) * 0.2*900 / 3600 / 1000
                       for t in range(T)) ,0,priority=2, abstol=0)"""

#optimization of CO2 emizzions of gas units
"""model.setObjectiveN(sum((b3_gas[t] +
                        b4_gas[t] +
                        b5_gas[t] +
                        b6_gas[t] +
                        b7_gas[t] +
                        b8_gas[t] +
                        b9_gas[t] +
                        b10_gas[t] +
                        b11_gas[t] +
                        b12_gas[t]) * 0.2*900 / 3600 / 1000
                       for t in range(T)) ,1,priority=1, abstol=0)"""

#optimization of variation of electricity
"""model.setObjectiveN(variationele,2,priority=7, abstol=0)"""

#optimization of variation of gas
"""model.setObjectiveN(variationgas,3,priority=6, abstol=0)"""

#dummy objective
"""model.setObjectiveN(1,1,priority=1, abstol=0)"""



model.optimize()
print()
wb = Workbook()
ws = wb.active

nSolutions = model.SolCount
nObjectives=6
dist=20
print('.................. ')
print('.................. ')
print('................... ')


for e in range(nSolutions):

    model.setParam(GRB.Param.SolutionNumber, e)
    ws.cell(row=nObjectives + dist - 12,  column=e + 2, value='variationele ')
    ws.cell(row=nObjectives + dist - 6, column=e + 2, value='variationgas ')
    ws.cell(row=nObjectives + dist + (T + 5) * 6, column=e + 2, value='co2pumpshedule')
    ws.cell(row=nObjectives + dist + (T + 5) * 28, column=e + 2, value='co2pumpgasshedule')
    ws.cell(row=nObjectives + dist + (T + 5) * 30, column=e + 2, value='EV charge   ')
    ws.cell(row=nObjectives + dist + (T + 5) * 3, column=e + 2, value='EV battery SOC')
    ws.cell(row=nObjectives + dist + (T + 5) * 31, column=e + 2, value='bus fill')
    ws.cell(row=nObjectives + dist + (T + 5) * 5, column=e + 2, value='bus tank SOC ')
    ws.cell(row=nObjectives + dist + (T + 5) * 1, column=e + 2, value='imported')
    ws.cell(row=nObjectives + dist + (T + 5) * 2, column=e + 2, value='gasSupplySchedule')
    ws.cell(row=nObjectives + dist + (T + 5) * 7, column=e + 2, value='batterySOC9 ')
    ws.cell(row=nObjectives + dist + (T + 5) * 8, column=e + 2, value='pumpSOC')
    ws.cell(row=nObjectives + dist + (T + 5) * 10, column=e + 2, value='waterSupplySchedule  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 11, column=e + 2, value='boilerSchedule3  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 12, column=e + 2, value='hpSchedule9  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 13, column=e + 2, value='batterySchedule9   ')
    ws.cell(row=nObjectives + dist + (T + 5) * 14, column=e + 2, value='chpDH_Schedule ')
    ws.cell(row=nObjectives + dist + (T + 5) * 15, column=e + 2, value='pumpSchedule ')
    ws.cell(row=nObjectives + dist + (T + 5) * 9, column=e + 2, value='utilization_pv  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 17, column=e + 2, value='utilization_wec  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 18, column=e + 2, value='b3_ele  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 19, column=e + 2, value='b3_gas  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 20, column=e + 2, value='b3_th   ')
    ws.cell(row=nObjectives + dist + (T + 5) * 21, column=e + 2, value='b3_wat   ')
    ws.cell(row=nObjectives + dist + (T + 5) * 22, column=e + 2, value='storageSOC  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 23, column=e + 2, value='storageSchedule    ')
    ws.cell(row=nObjectives + dist + (T + 5) * 24, column=e + 2, value='conversionSchedule    ')
    ws.cell(row=nObjectives + dist + (T + 5) * 25, column=e + 2, value='generalHP  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 26, column=e + 2, value='ngcfill  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 27, column=e + 2, value='ngctankSOC  ')
    ws.cell(row=nObjectives + dist + (T + 5) * 0 , column=e + 2, value='CO2 emissions  ')
    for r in range(27):
        ws.cell(row=123+(T+5)*r, column=e + 2, value='Total:  ')

    ws.cell(row=124 + (T + 5) * 0, column=e + 2, value=sum((imported[t].Xn/0.4+
                                                            chpDH_Schedule[t].Xn* dh_CHP.fNominal +
                                                            b3_gas[t].Xn*10-
                                                            co2max* 3 *co2pumpshedule[t].Xn
                                                            )*0.2*900 / 3600 / 1000 for t in range(T)))

    ws.cell(row=124 + (T + 5) * 1, column=e + 2, value=sum(imported[t].Xn * 900 / 3600 / 1000 for t in range(T)))

    ws.cell(row=124 + (T + 5) * 2, column=e + 2, value=sum((gasSupplySchedule[t].Xn * gs_nom)* 900 / 3600 / 1000 for t in range(T)))

    ws.cell(row=124 + (T + 5) * 3, column=e + 2, value=sum((EVbattSOC[t].Xn)/T for t in range(T)))



    """ws.cell(row=nObjectives + 8 + (T + 5) * 27, column=e + 2, value='  ')"""








    for t in range(T):
        ws.cell(row=nObjectives + dist - 11, column=e + 2, value=variationele.Xn)
        ws.cell(row=nObjectives + dist - 5, column=e + 2, value=variationgas.Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *6 + t, column=e + 2, value=co2pumpshedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *28 + t, column=e + 2, value=co2pumpgasshedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *30 + t, column=e + 2, value=EVcharge[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *3 + t, column=e + 2, value=EVbattSOC[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *31 + t, column=e + 2, value=busfill[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *5 + t, column=e + 2, value=bustankSOC[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *1 + t, column=e + 2, value=imported[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *2 + t, column=e + 2, value=gasSupplySchedule[t].Xn* gs_nom)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *7 + t, column=e + 2, value=batterySOC9[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *8 + t, column=e + 2, value=pumpSOC[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *10 + t, column=e + 2, value=waterSupplySchedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *11 + t, column=e + 2, value=boilerSchedule3[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *12 + t, column=e + 2, value=hpSchedule9[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *13 + t, column=e + 2, value=batterySchedule9[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *14 + t, column=e + 2, value=chpDH_Schedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *15 + t, column=e + 2, value=pumpSchedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) *9 + t, column=e + 2, value=utilization_pv[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 17 + t, column=e + 2, value=utilization_wec[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 18 + t, column=e + 2, value=b3_ele[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 19 + t, column=e + 2, value=b3_gas[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 20 + t, column=e + 2, value=b3_th[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 21 + t, column=e + 2, value=b3_wat[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 22 + t, column=e + 2, value=storageSOC[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 23 + t, column=e + 2, value=storageSchedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 24 + t, column=e + 2, value=conversionSchedule[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 25 + t, column=e + 2, value=generalHP[t].Xn)
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 26 + t, column=e + 2, value=ngcfill[t].Xn  )
        ws.cell(row=nObjectives + dist+1 + (T + 5) * 27 + t, column=e + 2, value=ngctankSOC[t].Xn  )
        ws.cell(row=nObjectives + dist + 1 + (T + 5) * 0 + t, column=e + 2, value=((imported[t].Xn/0.4+
                                                                                   chpDH_Schedule[t].Xn* dh_CHP.fNominal +
                                                                                   b3_gas[t].Xn*10-
                                                                                   co2max* 3 *co2pumpshedule[t].Xn
                                                                                   )*0.2*900 / 3600 / 1000))


        """ws.cell(row=nObjectives + 9 + (T + 5) * 27 + t, column=e + 2, value=[t].Xn)"""





nObjectives=model.NumObj

for e in range(nSolutions):
    print(' ')
    model.setParam(GRB.Param.SolutionNumber, e)
    print('optimization number %s' % e)
    print('------------------ ')

    for i in range(nObjectives):
        model.setParam(GRB.Param.ObjNumber, i)
        print('solution from objetive %s is %s .......' % (i, model.ObjNVal))
        ws.cell(row=i + 2, column=e + 2, value=model.ObjNVal)

print('.................. ')
print('.................. ')
print('................... ')
print(nSolutions)
print(nObjectives)
wb.save('optimization.xlsx')

"""lines modifed till here"""

"""if __name__ == '__main__':
    imported_Ele=0
    nonUtil_Ele=0
    directEmission=0
    for i in range(T):
        imported_Ele+=imported[i].X/1000/3600*900
        nonUtil_Ele+=((1 - utilization_wec[i].X) * cityDistrict.getWindEnergyConverterPower()[i] + (1 - utilization_pv[i].X) *cityDistrict.getPVPower()[i])/1000/3600*900
        directEmission+=(chpDH_Schedule[i].X * dh_CHP.fNominal +
                            b3_gas[i].X +
                            b4_gas[i].X +
                            b5_gas[i].X +
                            b6_gas[i].X +
                            b7_gas[i].X +
                            b8_gas[i].X +
                            b9_gas[i].X +
                            b10_gas[i].X +
                            b11_gas[i].X +
                            b12_gas[i].X) * 0.2/1000/3600*900
    print('import total',imported_Ele, 'kWh')
    print('non-utilized total',nonUtil_Ele,'kWh')
    print('direct emission',directEmission, 'kg CO2')"""

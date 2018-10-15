"""
author: ist-egu
"""

from openpyxl import Workbook
from openpyxl.styles import Font

from newFunctions.ModularCalculation_IO import *


def print2Excel(citydistrict, current_values=True):
    """
    Prints the simulation results to the excel sheets
    """
    wb = Workbook()
    wb.create_sheet('Electric')
    wb.create_sheet('Gas')
    wb.create_sheet('Heat')
    wb.create_sheet('Water')
    wb.create_sheet('Results')
    del (wb['Sheet'])
    ws_ele = wb['Electric']
    ws_gas = wb['Gas']
    ws_heat = wb['Heat']
    ws_water = wb['Water']
    ws_result = wb['Results']
    for sheet in wb.worksheets:
        sheet['A1'] = 'TimeStep'
        sheet['A1'].font = Font(bold=True)

    ele_gens = printEle(citydistrict,ws_ele, current_values=current_values)
    gas_gens = printGas(citydistrict,ws_gas, current_values=current_values)
    heat_gens = printThermal(citydistrict,ws_heat, current_values=current_values)
    water_gens = printWater(citydistrict,ws_water, current_values=current_values)
    printResult(citydistrict,ws_result, ele_gens, gas_gens, heat_gens, water_gens, current_values=current_values)

    return wb


def printResult(citydistrict, ws_res, ele, gas, heat, water, current_values=True):
    """
    Results sheet shows the resource balance and total emission
    """
    #TODO: remove citydistrict.ele_surplus
    emission = citydistrict.getEmission(current_values=current_values)

    if current_values:
        timesteps = citydistrict.environment.timer.timestepsUsedHorizon
    else:
        timesteps = citydistrict.environment.timer.timestepsTotal

    citydistrict.ele_surplus = np.zeros(timesteps)
    citydistrict.gas_surplus = np.zeros(timesteps)
    citydistrict.heat_surplus = np.zeros(timesteps)
    citydistrict.water_surplus = np.zeros(timesteps)

    for supplier in ele.keys():
        citydistrict.ele_surplus += ele[supplier]

    for supplier in gas.keys():
        citydistrict.gas_surplus += gas[supplier]

    for supplier in heat.keys():
        citydistrict.heat_surplus += heat[supplier]

    for supplier in water.keys():
        citydistrict.water_surplus += water[supplier]

    for t in range(timesteps):

        if t == 0:
            ws_res.cell(row=1, column=2, value='Total emission')
            ws_res.cell(row=1, column=3, value='Wasted electricty')
            ws_res.cell(row=1, column=4, value='Wasted gas')
            ws_res.cell(row=1, column=5, value='Wasted heat')
            ws_res.cell(row=1, column=6, value='Wasted water')
            ws_res._get_cell(1, 2).font = Font(bold=True)
            ws_res._get_cell(1, 3).font = Font(bold=True)
            ws_res._get_cell(1, 4).font = Font(bold=True)
            ws_res._get_cell(1, 5).font = Font(bold=True)
            ws_res._get_cell(1, 6).font = Font(bold=True)

        ws_res.cell(row=t + 2, column=1, value=str(t))
        ws_res.cell(row=t + 2, column=2, value=emission[t])
        ws_res.cell(row=t + 2, column=3, value=citydistrict.ele_surplus[t])
        ws_res.cell(row=t + 2, column=4, value=citydistrict.gas_surplus[t])
        ws_res.cell(row=t + 2, column=5, value=citydistrict.heat_surplus[t])
        ws_res.cell(row=t + 2, column=6, value=citydistrict.water_surplus[t])


def printEle(citydistrict, ws_ele, current_values=True):
    """
    Electricity sheet shows the impact of operation of buildings or devices at network nodes on the available electricity at the network
    -------
    supply power : np.array
        Array with net supply power values of all electrical objects
    """

    if current_values:
        timesteps = citydistrict.environment.timer.timestepsUsedHorizon
    else:
        timesteps = citydistrict.environment.timer.timestepsTotal

    # Create empty list of electrical entities
    ele_gen_schedules = {}
    electrical_storage = {}

    #  Loop over all nodes
    for n in citydistrict:
        #  If node holds attribute 'node_type'
        if 'node_type' in citydistrict.node[n]:
            #  If node_type is building
            if citydistrict.node[n]['node_type'] == 'building':
                #  If entity is of one of the following devices
                if citydistrict.node[n]['entity']._kind in ['pv', 'windenergyconverter', 'chp', 'heps', 'p2gConverter',
                                                    'p2gStorage', 'gassource', 'watersource']:
                    ele_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_device(citydistrict.node[n]['entity'], current_values)[0]
                    if citydistrict.node[n]['entity']._kind == 'heps':
                        electrical_storage[citydistrict.node[n]['entity']] = \
                        citydistrict.node[n]['entity'].getResults(currentValues=current_values)[0]
                # If entity is a building object
                if citydistrict.node[n]['entity']._kind == 'building':
                    ele_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_building(citydistrict.node[n]['entity'], current_values)[0]
                    if citydistrict.node[n]['entity'].hasBes and citydistrict.node[n]['entity'].bes.hasBattery:
                        electrical_storage[citydistrict.node[n]['entity'].bes.battery] = \
                        citydistrict.node[n]['entity'].bes.battery.getResults(currentValues=current_values)[0]

    if len(ele_gen_schedules) > 0 and len(electrical_storage):

        for t in range(timesteps):
            cnt = 2
            if t == 0:
                for building in sorted(ele_gen_schedules.keys(), key=lambda building: building.labels[0]):
                    ws_ele.cell(row=1, column=cnt, value=str(building.labels[0]) + '-' + building._kind)
                    ws_ele._get_cell(1, cnt).font = Font(bold=True)
                    cnt += 1

                for storage in sorted(electrical_storage.keys(), key=lambda stor: stor.labels[0]):
                    ws_ele.cell(row=1, column=cnt + 1, value=str(storage.labels[0]) + '-' + storage._kind)
                    ws_ele._get_cell(1, cnt + 1).font = Font(bold=True)
                    cnt += 1

                cnt = 2

            ws_ele.cell(row=t + 2, column=1, value=str(t))
            for building in sorted(ele_gen_schedules.keys(), key=lambda build: build.labels[0]):
                ws_ele.cell(row=t + 2, column=cnt, value=ele_gen_schedules[building][t])
                cnt += 1

            for storage in sorted(electrical_storage.keys(), key=lambda stor: stor.labels[0]):
                ws_ele.cell(row=t + 2, column=cnt + 1, value=electrical_storage[storage][t])
                cnt += 1

    return ele_gen_schedules


def printGas(citydistrict, ws_gas, current_values=True):
    """
    Gas sheet shows the impact of operation of buildings or devices at network nodes on the available gas at the network
    -------
    supply power : np.array
        Array with net supply power values of all gas objects
    """

    if current_values:
        timesteps = citydistrict.environment.timer.timestepsUsedHorizon
    else:
        timesteps = citydistrict.environment.timer.timestepsTotal

    # Create empty list of gas entities
    gas_gen_schedules = {}
    gas_storage = {}

    #  Loop over all nodes
    for n in citydistrict:
        #  If node holds attribute 'node_type'
        if 'node_type' in citydistrict.node[n]:
            #  If node_type is building
            if citydistrict.node[n]['node_type'] == 'building':
                #  If entity is of one of the following devices
                if citydistrict.node[n]['entity']._kind in ['pv', 'windenergyconverter', 'chp', 'heps', 'p2gConverter',
                                                    'p2gStorage', 'gassource', 'watersource']:
                    gas_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_device(citydistrict.node[n]['entity'], current_values)[1]
                    if citydistrict.node[n]['entity']._kind == 'p2gStorage':
                        gas_storage[citydistrict.node[n]['entity']] = \
                        citydistrict.node[n]['entity'].getResults(currentValues=current_values)[0]
                # If entity is a building object
                if citydistrict.node[n]['entity']._kind == 'building':
                    gas_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_building(citydistrict.node[n]['entity'], current_values)[1]

    if len(gas_gen_schedules) > 0:

        for t in range(timesteps):
            cnt = 2
            if t == 0:
                for building in sorted(gas_gen_schedules.keys(), key=lambda build: build.labels[0]):
                    ws_gas.cell(row=1, column=cnt, value=str(building.labels[0]) + '-' + building._kind)
                    ws_gas._get_cell(1, cnt).font = Font(bold=True)
                    cnt += 1
                for storage in sorted(gas_storage.keys(), key=lambda stor: stor.labels[0]):
                    ws_gas.cell(row=1, column=cnt + 1, value=str(storage.labels[0]) + '-' + storage._kind)
                    ws_gas._get_cell(1, cnt + 1).font = Font(bold=True)
                    cnt += 1

                cnt = 2

            ws_gas.cell(row=t + 2, column=1, value=str(t))
            for building in sorted(gas_gen_schedules.keys(), key=lambda build: build.labels[0]):
                ws_gas.cell(row=t + 2, column=cnt, value=gas_gen_schedules[building][t])
                cnt += 1

            for storage in sorted(gas_storage.keys(), key=lambda stor: stor.labels[0]):
                ws_gas.cell(row=t + 2, column=cnt + 1, value=gas_storage[storage][t])
                cnt += 1

    return gas_gen_schedules


def printThermal(citydistrict, ws_heat, current_values=True):
    """
    Heat sheet shows the impact of operation of buildings or devices at network nodes on the available heat at the DH network
    -------
    supply power : np.array
        Array with net supply power values of all gas objects
    """

    if current_values:
        timesteps = citydistrict.environment.timer.timestepsUsedHorizon
    else:
        timesteps = citydistrict.environment.timer.timestepsTotal

    # Create empty list of heat entities
    heat_gen_schedules = {}

    #  Loop over all nodes
    for n in citydistrict:
        #  If node holds attribute 'node_type'
        if 'node_type' in citydistrict.node[n]:
            #  If node_type is building
            if citydistrict.node[n]['node_type'] == 'building':
                #  If entity is of one of the following devices
                if citydistrict.node[n]['entity']._kind in ['pv', 'windenergyconverter', 'chp', 'heps', 'p2gConverter',
                                                    'p2gStorage', 'gassource', 'watersource']:
                    heat_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_device(citydistrict.node[n]['entity'], current_values)[2]
                # If entity is a building object
                if citydistrict.node[n]['entity']._kind == 'building':
                    heat_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_building(citydistrict.node[n]['entity'], current_values)[2]

    if len(heat_gen_schedules) > 0:

        for t in range(timesteps):
            cnt = 2
            if t == 0:
                for building in sorted(heat_gen_schedules.keys(), key=lambda build: build.labels[0]):
                    ws_heat.cell(row=1, column=cnt, value=str(building.labels[0]) + '-' + building._kind)
                    ws_heat._get_cell(1, cnt).font = Font(bold=True)
                    cnt += 1
                cnt = 2

            ws_heat.cell(row=t + 2, column=1, value=str(t))
            for building in sorted(heat_gen_schedules.keys(), key=lambda build: build.labels[0]):
                ws_heat.cell(row=t + 2, column=cnt, value=heat_gen_schedules[building][t])
                cnt += 1

    return heat_gen_schedules


def printWater(citydistrict, ws_water, current_values=True):
    """
    Water sheet shows the impact of operation of buildings or devices at network nodes on the available water at the water network
    -------
    supply water : np.array
        Array with net water volume of all electrical objects
    """

    if current_values:
        timesteps = citydistrict.environment.timer.timestepsUsedHorizon
    else:
        timesteps = citydistrict.environment.timer.timestepsTotal

    # Create empty list of pv entities
    water_gen_schedules = {}
    water_storage = {}

    #  Loop over all nodes
    for n in citydistrict:
        #  If node holds attribute 'node_type'
        if 'node_type' in citydistrict.node[n]:
            #  If node_type is building
            if citydistrict.node[n]['node_type'] == 'building':
                #  If entity is of one of the following devices
                if citydistrict.node[n]['entity']._kind in ['pv', 'windenergyconverter', 'chp', 'heps', 'p2gConverter',
                                                    'p2gStorage', 'gassource', 'watersource']:
                    water_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_device(citydistrict.node[n]['entity'], current_values)[3]
                    if citydistrict.node[n]['entity']._kind == 'heps':
                        water_storage[citydistrict.node[n]['entity']] = \
                        citydistrict.node[n]['entity'].getResults(currentValues=current_values)[0]
                # If entity is a building object
                if citydistrict.node[n]['entity']._kind == 'building':
                    water_gen_schedules[citydistrict.node[n]['entity']] = \
                    calculateIO_building(citydistrict.node[n]['entity'], current_values)[3]

    if len(water_gen_schedules) > 0:

        for t in range(timesteps):
            cnt = 2
            if t == 0:
                for building in sorted(water_gen_schedules.keys(), key=lambda build: build.labels[0]):
                    ws_water.cell(row=1, column=cnt, value=str(building.labels[0]) + '-' + building._kind)
                    ws_water._get_cell(1, cnt).font = Font(bold=True)
                    cnt += 1

                for storage in sorted(water_storage.keys(), key=lambda stor: stor.labels[0]):
                    ws_water.cell(row=1, column=cnt + 1, value=str(storage.labels[0]) + '-' + storage._kind)
                    ws_water._get_cell(1, cnt + 1).font = Font(bold=True)
                    cnt += 1

                cnt = 2

            ws_water.cell(row=t + 2, column=1, value=str(t))
            for building in sorted(water_gen_schedules.keys(), key=lambda build: build.labels[0]):
                ws_water.cell(row=t + 2, column=cnt, value=water_gen_schedules[building][t])
                cnt += 1

            for storage in sorted(water_storage.keys(), key=lambda stor: stor.labels[0]):
                ws_water.cell(row=t + 2, column=cnt + 1, value=water_storage[storage][t])
                cnt += 1

    return water_gen_schedules


